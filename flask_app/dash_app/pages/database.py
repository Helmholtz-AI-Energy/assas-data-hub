"""ASSAS Database Page.

This module provides the layout and functionality for the ASSAS Database page,
which allows users to view, search, and download datasets from the ASSAS training
dataset index.
"""

import os
import dash
import dash_bootstrap_components as dbc
import pandas as pd
import logging
import shutil
import math
import io
import base64

from pymongo import MongoClient
from dash import (
    dash_table,
    html,
    dcc,
    Input,
    Output,
    callback,
    State,
    callback_context,
)
from dash.exceptions import PreventUpdate
from flask import current_app as app
from zipfile import ZipFile
from uuid import uuid4
from pathlib import Path
from typing import List, Tuple

from assasdb import AssasDatabaseManager, AssasDatabaseHandler

# Define a common style dictionary
COMMON_STYLE = {
    "fontSize": "18px",
    "textAlign": "center",
    "margin": "1% auto",
    "padding": "10px",
    "width": "100%",
    "fontFamily": "arial, sans-serif",
}

# Define responsive layout using Bootstrap classes
RESPONSIVE_STYLE = {
    "width": "100%",
    "padding-left": "5%",
    "padding-right": "5%",
}

logger = logging.getLogger("assas_app")

colors = {"background": "#111111", "text": "#7FDBFF"}

operators = [
    ["ge ", ">="],
    ["le ", "<="],
    ["lt ", "<"],
    ["gt ", ">"],
    ["ne ", "!="],
    ["eq ", "="],
    ["contains "],
    ["datestartswith "],
]


def update_table_data() -> pd.DataFrame:
    """Update the table data from the database.

    Returns:
        pd.DataFrame: DataFrame with the table data.

    """
    logger.info("Load database entries to table.")

    database_manager = AssasDatabaseManager(
        database_handler=AssasDatabaseHandler(
            client=MongoClient(app.config["CONNECTIONSTRING"]),
            backup_directory=app.config["BACKUP_DIRECTORY"],
            database_name=app.config["MONGO_DB_NAME"],
        )
    )

    table_data_local = database_manager.get_all_database_entries()

    # 🔧 FIX: Clean complex data types that DataTable can't handle
    def clean_column_data(df):
        """Clean DataFrame columns to ensure DataTable compatibility."""
        df_clean = df.copy()
        
        # Convert complex data types to strings
        for col in df_clean.columns:
            if col in ['meta_data_variables', 'meta_keywords', 'meta_tags']:
                # Convert lists/dicts to JSON strings
                df_clean[col] = df_clean[col].apply(lambda x: 
                    str(x) if x is not None else ""
                )
            elif col.startswith('meta_') and df_clean[col].dtype == 'object':
                # Convert any other complex metadata to strings
                df_clean[col] = df_clean[col].astype(str).fillna("")
        
        return df_clean
    
    # Clean the data first
    table_data_local = clean_column_data(table_data_local)
    
    table_data_local["system_download"] = [
        f'<a href="/assas_app/hdf5_file?uuid={entry.system_uuid}">hdf5 file</a>'
        if entry.system_status == "Valid"
        else '<span class="no-download">no hdf5 file</span>'
        for entry in table_data_local.itertuples()
    ]
    table_data_local["meta_name"] = [
        f'<a href="/assas_app/details/{entry.system_uuid}">{entry.meta_name}</a>'
        for entry in table_data_local.itertuples()
    ]

    def get_status_html(
        status: str,
    ) -> str:
        status_classes = {
            "Valid": "status-valid",
            "Invalid": "status-invalid",
            "Converting": "status-converting",
            "Uploaded": "status-uploaded",
        }
        css_class = status_classes.get(status, "status-unknown")
        return f'<span class="{css_class}">{str(status)}</span>'

    table_data_local["system_status"] = [
        get_status_html(entry.system_status) for entry in table_data_local.itertuples()
    ]

    return table_data_local


def get_database_size() -> str:
    """Get the overall size of the ASSAS database.

    Returns:
        str: Overall size of the database in a human-readable format.

    """
    database_manager = AssasDatabaseManager(
        database_handler=AssasDatabaseHandler(
            client=MongoClient(app.config["CONNECTIONSTRING"]),
            backup_directory=app.config["BACKUP_DIRECTORY"],
            database_name=app.config["MONGO_DB_NAME"],
        )
    )
    size = database_manager.get_overall_database_size()

    if size is None or len(size) == 0:
        return "0 B"

    return size


table_data = update_table_data()
database_size = get_database_size()
now = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")

ALL = len(table_data)
PAGE_SIZE = 30
PAGE_MAX_SIZE = 100

PAGE_COUNT = ALL / PAGE_SIZE

dash.register_page(__name__, path="/database")


# Updated modern style dictionary
MODERN_STYLE = {
    "fontFamily": "Arial, sans-serif",
    "fontSize": "16px",
    "color": "#2c3e50",
    "lineHeight": "1.6",
}

CARD_STYLE = {
    "backgroundColor": "#ffffff",
    "border": "1px solid #e0e6ed",
    "borderRadius": "8px",
    "boxShadow": "0 2px 8px rgba(0, 0, 0, 0.1)",
    "padding": "1.5rem",
    "marginBottom": "1.5rem",
    "transition": "all 0.3s ease",
}

BUTTON_STYLE = {
    "fontFamily": "Arial, sans-serif",
    "fontWeight": "600",
    "borderRadius": "6px",
    "padding": "12px 24px",
    "fontSize": "14px",
    "transition": "all 0.3s ease",
    "border": "none",
    "cursor": "pointer",
}


def layout() -> html.Div:
    """Layout for the ASSAS Database page - Mobile Responsive.

    Returns:
        html.Div: The layout of the ASSAS Database page.

    """
    return_div = \
        html.Div(
        [
        # Header Section - Mobile Responsive
        dbc.Container(
        [
        dbc.Row(
        [
        dbc.Col(
        [
        html.H1(
        "ASSAS Database Training Dataset Index",
        className="display-6 fw-bold text-primary mb-3 text-center",
        # Changed from display-4 to display-6 for mobile
        style={
        "fontSize": "clamp(1.5rem, 4vw, 3rem)",  # Responsive font size
        "lineHeight": "1.2",
        },
        ),
        html.H2(
        "Browse, Search, and Download Machine Learning Training Datasets",
        className="text-secondary mb-4 text-center",
        style={
        "fontSize": "clamp(0.9rem, 2.5vw, 1.5rem)",  # Responsive font size
        "fontWeight": "400",
        "lineHeight": "1.4",
        },
        ),
        html.Div(
        [
        dbc.Button(
        [
        html.I(
        className="fas fa-question-circle me-2"
        ),
        html.Span(
        "Usage Guide",
        className="d-none d-sm-inline",
        ),  # Hide text on very small screens
        html.Span(
        "Guide",
        className="d-inline d-sm-none",
        ),  # Show short text on small screens
        ],
        id="toggle-usage-guide",
        color="primary",
        outline=False,
        size="lg",
        className="w-100 mb-3",
        style={
        **BUTTON_STYLE,
        "backgroundColor": "#17a2b8",
        "borderColor": "#17a2b8",
        "color": "#ffffff",
        "fontSize": "clamp(0.8rem, 2vw, 1rem)",  # Responsive button text
        },
        ),
        dbc.Collapse(
        [
        dbc.Card(
        [
        dbc.CardBody(
        [
        html.H4(
        "How to Use This Database Interface",
        className="text-primary mb-4",
        style={
        "fontSize": "clamp(1rem, 3vw, 1.4rem)",
        "fontWeight": "600",
        },
        ),
        # Mobile-responsive grid with expanded content
        dbc.Row(
        [
        # Searching & Filtering Section
        dbc.Col(
        [
        html.H5(
        "🔍 Searching & Filtering",
        className="text-secondary mb-3",
        style={
        "fontSize": "clamp(0.9rem, 2.5vw, 1.1rem)"
        },
        ),
        html.Ul(
        [
        html.Li(
        [
        html.Strong(
        "Sort columns: "
        ),
        "Click on column headers to sort data ascending/descending",
        ],
        className="mb-2",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        html.Li(
        [
        html.Strong(
        "Filter data: "
        ),
        "Use filter boxes that appear when hovering over column headers",
        ],
        className="mb-2",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        html.Li(
        [
        html.Strong(
        "Multi-column sorting: "
        ),
        "Hold Shift and click multiple headers for complex sorting",
        ],
        className="mb-2",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        html.Li(
        [
        html.Strong(
        "Clear filters: "
        ),
        "Delete text in filter boxes to remove filters",
        ],
        className="mb-2",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        ]
        ),
        ],
        xs=12,
        lg=4,
        className="mb-4",
        ),
        # Filter Examples Section
        dbc.Col(
        [
        html.H5(
        "📝 Filter Examples",
        className="text-secondary mb-3",
        style={
        "fontSize": "clamp(0.9rem, 2.5vw, 1.1rem)"
        },
        ),
        html.Div(
        [
        html.P(
        [
        html.Strong(
        "Text filtering:"
        ),
        ],
        className="mb-1",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        html.Ul(
        [
        html.Li(
        [
        html.Code(
        "Valid",
        className="bg-light px-1",
        ),
        " - Show only Valid datasets",
        ],
        className="mb-1",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        html.Li(
        [
        html.Code(
        "contains example",
        className="bg-light px-1",
        ),
        " - Find datasets containing 'example'",
        ],
        className="mb-1",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        html.Li(
        [
        html.Code(
        "ne Invalid",
        className="bg-light px-1",
        ),
        " - Exclude Invalid datasets",
        ],
        className="mb-2",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        ]
        ),
        html.P(
        [
        html.Strong(
        "Date filtering:"
        ),
        ],
        className="mb-1",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        html.Ul(
        [
        html.Li(
        [
        html.Code(
        "datestartswith 2024",
        className="bg-light px-1",
        ),
        " - Show 2024 datasets",
        ],
        className="mb-1",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        html.Li(
        [
        html.Code(
        ">= 2024-01-01",
        className="bg-light px-1",
        ),
        " - From January 2024 onwards",
        ],
        className="mb-2",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        ]
        ),
        html.P(
        [
        html.Strong(
        "Size filtering:"
        ),
        ],
        className="mb-1",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        html.Ul(
        [
        html.Li(
        [
        html.Code(
        "> 100",
        className="bg-light px-1",
        ),
        " - Larger than 100 MB",
        ],
        className="mb-1",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        html.Li(
        [
        html.Code(
        "< 1000",
        className="bg-light px-1",
        ),
        " - Smaller than 1 GB",
        ],
        className="mb-1",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        ]
        ),
        ]
        ),
        ],
        xs=12,
        lg=4,
        className="mb-4",
        ),
        # Navigation & Downloads Section
        dbc.Col(
        [
        html.H5(
        "📄 Navigation & Downloads",
        className="text-secondary mb-3",
        style={
        "fontSize": "clamp(0.9rem, 2.5vw, 1.1rem)"
        },
        ),
        html.Ul(
        [
        html.Li(
        [
        html.Strong(
        "Navigate pages: "
        ),
        "Use pagination controls at the bottom",
        ],
        className="mb-2",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        html.Li(
        [
        html.Strong(
        "Page size: "
        ),
        "Change entries per page (1-100 datasets)",
        ],
        className="mb-2",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        html.Li(
        [
        html.Strong(
        "Dataset details: "
        ),
        "Click on dataset names for detailed information",
        ],
        className="mb-2",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        html.Li(
        [
        html.Strong(
        "Single download: "
        ),
        "Click 'hdf5 file' links in Download column",
        ],
        className="mb-2",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        ]
        ),
        ],
        xs=12,
        lg=4,
        className="mb-4",
        ),
        ]
        ),
        # Bulk Download Section
        dbc.Row(
        [
        dbc.Col(
        [
        html.Hr(
        className="my-4"
        ),
        html.H5(
        "📦 Bulk Download Instructions",
        className="text-primary mb-3",
        style={
        "fontSize": "clamp(0.9rem, 2.5vw, 1.1rem)"
        },
        ),
        dbc.Row(
        [
        dbc.Col(
        [
        html.H6(
        "Step 1: Select Datasets",
        className="text-secondary mb-2",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        html.Ul(
        [
        html.Li(
        "Click the checkbox at the beginning of each row",
        className="mb-1",
        style={
            "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        html.Li(
        "Select multiple datasets (up to 20 at once)",
        className="mb-1",
        style={
            "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        html.Li(
        "Use filters first to narrow down your selection",
        className="mb-2",
        style={
            "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        ]
        ),
        ],
        xs=12,
        md=6,
        className="mb-3",
        ),
        dbc.Col(
        [
        html.H6(
        "Step 2: Generate Download",
        className="text-secondary mb-2",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        html.Ul(
        [
        html.Li(
        "Open 'Database Tools' section above",
        className="mb-1",
        style={
            "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        html.Li(
        "Click 'Get Download Link' button",
        className="mb-1",
        style={
            "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        html.Li(
        "Wait for ZIP archive to be prepared",
        className="mb-2",
        style={
            "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        ]
        ),
        ],
        xs=12,
        md=6,
        className="mb-3",
        ),
        ]
        ),
        dbc.Alert(
        [
        html.I(
        className="fas fa-info-circle me-2"
        ),
        html.Strong(
        "Download Limits: "
        ),
        "Maximum 20 datasets per download. Only 'Valid' datasets include HDF5 files. "
        "Invalid datasets show 'no hdf5 file'.",
        ],
        color="info",
        className="mb-3",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        ],
        xs=12,
        )
        ]
        ),
        # Advanced Tips Section
        dbc.Row(
        [
        dbc.Col(
        [
        html.Hr(
        className="my-4"
        ),
        html.H5(
        "💡 Advanced Tips",
        className="text-primary mb-3",
        style={
        "fontSize": "clamp(0.9rem, 2.5vw, 1.1rem)"
        },
        ),
        dbc.Row(
        [
        dbc.Col(
        [
        html.H6(
        "Efficient Filtering",
        className="text-secondary mb-2",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        html.Ul(
        [
        html.Li(
        [
        "Combine filters: Use ",
        html.Code(
        "Status = Valid",
        className="bg-light px-1",
        ),
        " and date filters together",
        ],
        className="mb-1",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        html.Li(
        "Filter before selecting: Reduces scrolling through pages",
        className="mb-1",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        html.Li(
        "Use wildcards: Text filters are case-insensitive",
        className="mb-2",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        ]
        ),
        ],
        xs=12,
        md=6,
        className="mb-3",
        ),
        dbc.Col(
        [
        html.H6(
        "Download Strategy",
        className="text-secondary mb-2",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        html.Ul(
        [
        html.Li(
        "Check file sizes before bulk download",
        className="mb-1",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        html.Li(
        "Download in batches for large collections",
        className="mb-1",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        html.Li(
        "Valid datasets have both binary and HDF5 files",
        className="mb-2",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        ]
        ),
        ],
        xs=12,
        md=6,
        className="mb-3",
        ),
        ]
        ),
        ],
        xs=12,
        )
        ]
        ),
        # Final Pro Tips
        dbc.Row(
        [
        dbc.Col(
        [
        dbc.Alert(
        [
        html.I(
        className="fas fa-lightbulb me-2"
        ),
        html.Strong(
        "Pro Tips: "
        ),
        html.Ul(
        [
        html.Li(
        "Use browser's back button to return to filtered results",
        className="mb-1",
        ),
        html.Li(
        "Bookmark filtered URLs for repeated searches",
        className="mb-1",
        ),
        html.Li(
        "Check dataset details page for metadata and preview",
        className="mb-1",
        ),
        html.Li(
        "Refresh page to get latest data from database",
        className="mb-0",
        ),
        ],
        className="mb-0 mt-2",
        ),
        ],
        color="success",
        className="mb-0",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        )
        ],
        xs=12,
        )
        ]
        ),
        ]
        )
        ],
        style={
        **CARD_STYLE,
        "backgroundColor": "#f8f9fa",
        "border": "1px solid #dee2e6",
        },
        )
        ],
        id="collapse-usage-guide",
        is_open=False,
        ),
        ],
        className="mb-4",
        ),
        ],
        xs=12,
        )  # Full width on all screen sizes
        ]
        )
        ],
        fluid=True,
        className="mb-4",
        ),
        # Database Tools Section - Mobile Responsive
        dbc.Container(
        [
        dbc.Row(
        [
        dbc.Col(
        [
        dbc.Button(
        [
        html.I(className="fas fa-tools me-2"),
        html.Span(
        "Database Tools",
        className="d-none d-sm-inline",
        ),
        html.Span(
        "Tools", className="d-inline d-sm-none"
        ),
        ],
        id="toggle-section",
        color="primary",
        size="lg",
        className="w-100 mb-3",
        style={
        **BUTTON_STYLE,
        "fontSize": "clamp(0.8rem, 2vw, 1rem)",
        },
        ),
        dbc.Collapse(
        [
        dbc.Card(
        [
        dbc.CardBody(
        [
        # Mobile-stacked layout
        dbc.Row(
        [
        # Download Section
        dbc.Col(
        [
        html.H5(
        [
        html.I(
        className="fas fa-download me-2 text-primary"
        ),
        "Download",
        ],
        className="mb-3",
        style={
        "fontSize": "clamp(0.9rem, 2.5vw, 1.1rem)"
        },
        ),
        dcc.Loading(
        [
        dbc.Button(
        [
        html.I(
        className="fas fa-file-archive me-2"
        ),
        html.Span(
        "Get Download Link",
        className="d-none d-md-inline",
        ),
        html.Span(
        "Download",
        className="d-inline d-md-none",
        ),
        ],
        id="download_selected",
        color="success",
        size="lg",
        disabled=True,
        className="w-100 mb-3",
        style={
        **BUTTON_STYLE,
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)",
        },
        )
        ],
        type="default",
        ),
        html.Div(
        "Select datasets to enable download",
        id="download_selected_info",
        className="text-muted small text-center mb-2",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        dbc.Alert(
        [
        html.I(
        className="fas fa-exclamation-triangle me-2"
        ),
        "Max 20 datasets",
        ],
        color="warning",
        className="small",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        ],
        xs=12,
        md=4,
        className="mb-3 mb-md-0",
        ),  # Stack on mobile
        # Refresh Section
        dbc.Col(
        [
        html.H5(
        [
        html.I(
        className="fas fa-sync-alt me-2 text-primary"
        ),
        "Refresh",
        ],
        className="mb-3",
        style={
        "fontSize": "clamp(0.9rem, 2.5vw, 1.1rem)"
        },
        ),
        dbc.Button(
        [
        html.I(
        className="fas fa-refresh me-2"
        ),
        "Refresh",
        ],
        id="reload_page",
        color="info",
        size="lg",
        className="w-100 mb-3",
        style={
        **BUTTON_STYLE,
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)",
        },
        ),
        html.Div(
        [
        html.Strong(
        "Datasets: "
        ),
        html.Span(
        f"{len(table_data)}",
        className="text-primary",
        ),
        ],
        id="dataset_count",
        className="text-center mb-2",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        html.Div(
        f"Updated: {now[:16]}",  # Shortened timestamp for mobile
        id="database_update_time",
        className="text-muted small text-center",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        ],
        xs=12,
        md=4,
        className="mb-3 mb-md-0",
        ),
        # Storage Info Section
        dbc.Col(
        [
        html.H5(
        [
        html.I(
        className="fas fa-hdd me-2 text-primary"
        ),
        "Storage",
        ],
        className="mb-3",
        style={
        "fontSize": "clamp(0.9rem, 2.5vw, 1.1rem)"
        },
        ),
        dbc.Table(
        [
        html.Tbody(
        [
        html.Tr(
        [
        html.Td(
        "Used",
        className="fw-bold",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        html.Td(
        database_size,
        className="text-end",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        ]
        ),
        html.Tr(
        [
        html.Td(
        "Limit",
        className="fw-bold",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        html.Td(
        "100 TB",
        className="text-end",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        ]
        ),
        ]
        )
        ],
        size="sm",
        striped=True,
        hover=True,
        className="mb-0",
        ),
        ],
        xs=12,
        md=4,
        ),
        ]
        ),
        # Download Link Section
        dbc.Row(
        [
        dbc.Col(
        [
        html.Hr(),
        html.Div(
        [
        html.H6(
        "Download Link:",
        className="mb-2",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        html.Div(
        "Download link will appear here",
        id="download_link",
        className="text-center p-3 bg-light rounded",
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        ]
        ),
        ],
        xs=12,
        )
        ]
        ),
        ]
        )
        ],
        style=CARD_STYLE,
        )
        ],
        id="collapse-section",
        is_open=False,
        ),
        ],
        xs=12,
        )
        ]
        )
        ],
        fluid=True,
        className="mb-4",
        ),
        # Data Table Section - Mobile Responsive
        dbc.Container(
        [
        dbc.Row(
        [
        dbc.Col(
        [
        dbc.Card(
        [
        dbc.CardHeader(
        [
        # Enhanced header with export buttons
        dbc.Row(
        [
        # Title Section (Left)
        dbc.Col(
        [
        html.H4(
        [
        html.I(className="fas fa-table me-2"),
        html.Span(
        "Dataset Overview",
        className="d-none d-sm-inline",
        ),
        html.Span(
        "Datasets",
        className="d-inline d-sm-none",
        ),
        ],
        className="mb-0 text-primary",
        style={
        "fontSize": "clamp(1rem, 3vw, 1.5rem)"
        },
        )
        ],
        xs=12,
        md=6,
        className="d-flex align-items-center mb-2 mb-md-0"
        ),
        # Export Buttons Section (Right)
        dbc.Col(
        [
        html.Div(
        [
        # Export label (hidden on mobile)
        html.Span(
        "Export:",
        className="me-2 text-muted fw-bold d-none d-lg-inline",
        style={
        "fontSize": "0.9rem",
        "alignSelf": "center"
        }
        ),
        # Export buttons group
        dbc.ButtonGroup(
        [
        dbc.Button(
        [
        html.I(className="fas fa-file-csv me-1"),
        html.Span("CSV", className="d-none d-md-inline"),
        ],
        id="export-csv-btn",
        color="success",
        outline=True,
        size="sm",
        className="export-btn",
        style={
        "fontSize": "clamp(0.7rem, 1.5vw, 0.8rem)",
        "fontWeight": "600",
        "borderRadius": "0.375rem 0 0 0.375rem",
        "minWidth": "50px"
        }
        ),
        dbc.Button(
        [
        html.I(className="fas fa-file-excel me-1"),
        html.Span("Excel", className="d-none d-md-inline"),
        ],
        id="export-excel-btn", 
        color="success",
        outline=True,
        size="sm",
        className="export-btn",
        style={
        "fontSize": "clamp(0.7rem, 1.5vw, 0.8rem)",
        "fontWeight": "600",
        "borderRadius": "0 0.375rem 0.375rem 0",
        "minWidth": "50px"
        }
        )
        ],
        size="sm"
        ),
        ],
        className="d-flex align-items-center justify-content-end"
        ),
        # Export status
        html.Div(
        "",
        id="export-status-database",
        className="text-muted small text-end mt-1",
        style={
        "fontSize": "0.7rem",
        "minHeight": "1rem"
        }
        )
        ],
        xs=12,
        md=6,
        className="d-flex flex-column align-items-end"
        ),
        ],
        className="align-items-center",
        )
        ],
        style={
        "padding": "1rem",
        "backgroundColor": "#f8f9fa",
        },
        ),
        # Add export options BELOW the settings section in CardBody:
        dbc.CardBody(
        [
        # Mobile-responsive pagination and settings (existing code)
        dbc.Row(
        [
        dbc.Col(
        [
        html.Div(
        [
        html.H6(
        "Navigation",
        className="mb-2 text-secondary",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        # Mobile-first pagination with conditional visibility
        html.Div(
        [
        # Mobile pagination (simple prev/next + current page)
        dbc.ButtonGroup(
        [
        dbc.Button(
        [
        html.I(
        className="fas fa-chevron-left"
        )
        ],
        id="mobile-prev-btn",
        size="sm",
        outline=True,
        color="primary",
        style={
        "minWidth": "40px",
        "fontSize": "0.8rem",
        },
        ),
        dbc.Button(
        [
        html.Span(
        "1",
        id="mobile-current-page",
        )
        ],
        size="sm",
        disabled=True,
        color="secondary",
        style={
        "minWidth": "50px",
        "fontSize": "0.8rem",
        },
        ),
        dbc.Button(
        [
        html.I(
        className="fas fa-chevron-right"
        )
        ],
        id="mobile-next-btn",
        size="sm",
        outline=True,
        color="primary",
        style={
        "minWidth": "40px",
        "fontSize": "0.8rem",
        },
        ),
        ],
        className="d-flex d-md-none justify-content-center mb-2",
        ),  # Show only on mobile
        # Desktop pagination (full pagination)
        dbc.Pagination(
        id="pagination",
        first_last=True,
        previous_next=True,
        max_value=int(
        PAGE_COUNT
        ),
        fully_expanded=False,
        size="sm",  # Use small size for better mobile compatibility
        className="justify-content-center mb-2 d-none d-md-flex",
        # Hide on mobile, show on desktop
        ),
        ]
        ),
        html.Small(
        [
        html.I(
        className="fas fa-info-circle me-1"
        ),
        html.Span(
        "Page 1/10",
        id="pagination-contents",
        ),
        ],
        className="text-muted text-center d-block",
        style={
        "fontSize": "clamp(0.6rem, 1.5vw, 0.7rem)"
        },
        ),
        ],
        style={
        "textAlign": "center"
        },
        )
        ],
        xs=12,
        md=8,
        className="mb-3 mb-md-0",
        ),
        dbc.Col(
        [
        html.Div(
        [
        html.H6(
        "Settings",
        className="mb-3 text-secondary",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        dbc.InputGroup(
        [
        dbc.InputGroupText(
        [
        html.I(
        className="fas fa-list-ol me-1"
        ),
        html.Span(
        "Datasets per Page:",
        className="d-none d-lg-inline",
        ),
        html.Span(
        "Size:",
        className="d-inline d-lg-none",
        ),
        ],
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)"
        },
        ),
        dbc.Input(
        id="datatable-page-size",
        type="number",
        min=1,
        max=PAGE_MAX_SIZE,
        value=PAGE_SIZE,
        placeholder=str(
        PAGE_SIZE
        ),
        disabled=False,
        style={
        "fontSize": "clamp(0.7rem, 1.8vw, 0.8rem)",
        "width": "80px",
        "maxWidth": "80px",
        "minWidth": "60px",
        "textAlign": "center",
        "flex": "0 0 auto",
        },
        ),
        ],
        className="mb-3 responsive-input-group",
        ),
        ],
        style={
        "position": "relative"
        },
        )
        ],
        xs=12,
        md=4,
        ),
        ],
        className="mb-3",
        ),
        # NEW: Export Options Section (below settings)
        dbc.Row(
        [
        dbc.Col(
        [
        html.Div(
        [
        html.H6(
        [
        html.I(className="fas fa-file-export me-2"),
        "Export Options"
        ],
        className="mb-2 text-secondary",
        style={
        "fontSize": "clamp(0.8rem, 2vw, 0.9rem)"
        },
        ),
        dbc.Checklist(
        options=[
        {"label": "Current page only", "value": "current_page"},
        {"label": "Apply current filters", "value": "apply_filters"},
        {"label": "Include metadata links", "value": "include_links"}
        ],
        value=["apply_filters"],
        id="export-options",
        inline=True,
        className="export-options-list",
        style={
        "fontSize": "clamp(0.7rem, 1.5vw, 0.75rem)",
        "color": "#6c757d"
        }
        )
        ],
        className="export-options-container"
        )
        ],
        xs=12,
        className="mb-3"
        )
        ]
        ),
        # Enhanced Table Container - Mobile Responsive (existing table code stays the same)
        html.Div(
        [
        dash_table.DataTable(
        id="datatable-paging-and-sorting",
        columns=[
        {
        "name": "Index",
        "id": "system_index",
        "selectable": True,
        "type": "numeric",
        "deletable": False,
        "renamable": False,
        "hideable": False,
        },
        {
        "name": "Dataset Name",
        "id": "meta_name",
        "selectable": True,
        "presentation": "markdown",
        "type": "text",
        "deletable": False,
        "renamable": False,
        "hideable": False,
        },
        {
        "name": "Status",
        "id": "system_status",
        "selectable": True,
        "presentation": "markdown",
        "type": "text",
        "deletable": False,
        "renamable": False,
        "hideable": False,
        },
        {
        "name": "Date",
        "id": "system_date",
        "selectable": True,
        "type": "datetime",
        "deletable": False,
        "renamable": False,
        "hideable": False,
        },
        {
        "name": "User",
        "id": "system_user",
        "selectable": True,
        "type": "text",
        "deletable": False,
        "renamable": False,
        "hideable": False,
        },
        {
        "name": "Binary Size",
        "id": "system_size",
        "selectable": True,
        "type": "text",
        "deletable": False,
        "renamable": False,
        "hideable": False,
        },
        {
        "name": "HDF5 Size",
        "id": "system_size_hdf5",
        "selectable": True,
        "type": "text",
        "deletable": False,
        "renamable": False,
        "hideable": False,
        },
        {
        "name": "Download",
        "id": "system_download",
        "selectable": True,
        "presentation": "markdown",
        "type": "text",
        "deletable": False,
        "renamable": False,
        "hideable": False,
        },
        ],
        data=[],
        # ENABLE SORTING AND FILTERING
        sort_action="custom",
        sort_mode="multi",
        filter_action="custom",
        filter_query="",
        # PAGINATION SETTINGS
        page_current=0,
        page_size=PAGE_SIZE,
        page_action="none",
        # SELECTION SETTINGS
        row_selectable="multi",
        selected_rows=[],
        # MOBILE-RESPONSIVE STYLING
        style_table={
        "width": "100%",
        "height": "auto",
        "overflowX": "auto",  # Horizontal scroll on mobile
        "overflowY": "visible",
        "minWidth": "800px",  # Ensure minimum width for all columns
        },
        style_header={
        "backgroundColor": "#007bff",
        "color": "white",
        "fontWeight": "bold",
        "textAlign": "center",
        "border": "1px solid #dee2e6",
        "fontSize": "clamp(10px, 2vw, 14px)",  # Responsive font
        "padding": "8px 4px",  # Smaller padding for mobile
        "whiteSpace": "nowrap",
        },
        style_cell={
        "textAlign": "center",
        "padding": "6px 4px",  # Smaller padding
        "fontFamily": "Arial, sans-serif",
        "fontSize": "clamp(10px, 1.8vw, 13px)",  # Responsive font
        "border": "1px solid #dee2e6",
        "whiteSpace": "nowrap",  # Prevent text wrapping
        "overflow": "hidden",
        "textOverflow": "ellipsis",
        "verticalAlign": "middle",
        "minHeight": "40px",
        },
        style_data={
        "backgroundColor": "white",
        "color": "black",
        "border": "1px solid #dee2e6",
        },
        style_data_conditional=[
        {
        "if": {
        "row_index": "odd"
        },
        "backgroundColor": "#f8f9fa",
        },
        {
        "if": {
        "state": "selected"
        },
        "backgroundColor": "#d1ecf1",
        "border": "1px solid #bee5eb",
        },
        ],
        style_cell_conditional=[
        {
        "if": {
        "column_id": "system_index"
        },
        "width": "60px",
        "minWidth": "60px",
        "maxWidth": "60px",
        },
        {
        "if": {
        "column_id": "meta_name"
        },
        "width": "200px",
        "minWidth": "150px",
        "maxWidth": "250px",
        "whiteSpace": "normal",
        "textAlign": "left",
        },
        {
        "if": {
        "column_id": "system_status"
        },
        "width": "80px",
        "minWidth": "80px",
        "maxWidth": "100px",
        },
        {
        "if": {
        "column_id": "system_date"
        },
        "width": "100px",
        "minWidth": "100px",
        "maxWidth": "120px",
        },
        {
        "if": {
        "column_id": "system_user"
        },
        "width": "80px",
        "minWidth": "80px",
        "maxWidth": "100px",
        },
        {
        "if": {
        "column_id": "system_size"
        },
        "width": "80px",
        "minWidth": "80px",
        "maxWidth": "100px",
        },
        {
        "if": {
        "column_id": "system_size_hdf5"
        },
        "width": "80px",
        "minWidth": "80px",
        "maxWidth": "100px",
        },
        {
        "if": {
        "column_id": "system_download"
        },
        "width": "120px",
        "minWidth": "120px",
        "maxWidth": "150px",
        },
        ],
        # ADDITIONAL OPTIONS
        merge_duplicate_headers=True,
        markdown_options={
        "html": True
        },
        css=[
        {
        "selector": ".dash-table-container",
        "rule": "overflow-x: auto !important;",
        },
        {
        "selector": ".dash-table-container .dash-spreadsheet-container",
        "rule": "overflow-x: auto !important;",
        },
        {
        "selector": ".dash-table-container table",
        "rule": "min-width: 800px !important;",
        },
        ],
        )
        ],
        className="table-responsive",
        style={
        "overflow": "auto",
        "minHeight": "300px",
        "height": "auto",
        "maxHeight": "none",
        "padding": "0.5rem",
        "backgroundColor": "#ffffff",
        "borderRadius": "8px",
        "boxShadow": "inset 0 2px 4px rgba(0, 0, 0, 0.05)",
        },
        ),
        # Hidden download components
        dcc.Download(id="download-csv-database"),
        dcc.Download(id="download-excel-database"),
        ],
        style={"padding": "1rem"},
        ),
        ],
        style={
        **CARD_STYLE,
        "minHeight": "auto",
        "height": "auto",
        "width": "100%",
        "maxWidth": "100%",
        "boxShadow": "0 4px 12px rgba(0, 0, 0, 0.1)",  # Smaller shadow
        "border": "1px solid #e0e6ed",
        },
        )
        ],
        xs=12,
        className="mb-4",
        )
        ]
        )
        ],
        fluid=True,
        className="mb-4",
        ),
        # Hidden elements
        dcc.Location(id="download_location", refresh=True),
        ],
        style={
        "backgroundColor": "#f8f9fa",
        "minHeight": "100vh",
        "paddingTop": "1rem",  # Reduced padding for mobile
        "paddingBottom": "1rem",
        },
        )

    return return_div

@callback(
    Output("collapse-section", "is_open"),
    Input("toggle-section", "n_clicks"),
    prevent_initial_call=True,
)
def toggle_section(n_clicks: int) -> bool:
    """Toggle the visibility of the database tools section.

    Args:
        n_clicks (int): Number of clicks on the toggle button.

    Returns:
        bool: True if the section should be open, False if it should be closed.

    """
    logger.info(f"Toggle section clicked {n_clicks} times.")
    return n_clicks % 2 == 1  # Toggle between open and closed


def copy_and_zip_files(
    file_info: List[tuple], destination_folder: str, zip_file_name: str
) -> str:
    """Copy a list of files to a destination folder and zips them into an archive.

    Args:
        file_info (List[tuple]): List of tuples containing file paths and UUIDs.
            Each tuple should be in the format (file_path, file_uuid).
        destination_folder (str): Path to the folder where files will be copied.
        zip_file_name (str): Name of the zip file to be created.

    Returns:
        str: Path to the created zip file.

    """
    # Ensure the destination folder exists
    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    # Copy files to the destination folder
    for file_path, file_uuid in file_info:
        if os.path.exists(file_path):
            destination_path = Path(destination_folder) / (
                f"{file_uuid}_" + Path(file_path).name
            )
            logger.info(f"Copying {file_path} to {destination_path}")
            # Copy the file to the destination folder with a new name
            shutil.copy(file_path, destination_path)
        else:
            logger.warning(f"File not found: {file_path}")

    file_list = [
        os.path.join(destination_folder, f"{file_uuid}_" + Path(file_path).name)
        for file_path, file_uuid in file_info
    ]
    logger.info(f"Files copied to {destination_folder}: {file_list}")

    # Create the zip file
    zip_file_path = os.path.join(destination_folder, zip_file_name)
    with ZipFile(zip_file_path, "w") as zip_object:
        for file_path in file_list:
            file_name = Path(file_path).name
            destination_path = os.path.join(destination_folder, file_name)
            if os.path.exists(destination_path):
                zip_object.write(destination_path, arcname=file_name)

    # Verify if the zip file was created
    if os.path.exists(zip_file_path):
        logger.info(f"ZIP file created: {zip_file_path}")
        return zip_file_path
    else:
        logger.error(f"Failed to create ZIP file: {zip_file_path}")
        return None


def clean_tmp_folder(parent_folder: str) -> None:
    """Delete all folders within a specified parent folder.

    Args:
        parent_folder (str): Path to the parent folder.

    """
    try:
        parent_path = Path(parent_folder)
        if not parent_path.exists():
            logger.info(f"Parent folder {parent_folder} does not exist.")
            return

        # Iterate through all items in the parent folder
        for item in parent_path.iterdir():
            if item.is_dir():  # Check if the item is a folder
                logger.info(f"Deleting folder: {item}")
                shutil.rmtree(item)  # Delete the folder and its contents

        logger.info(f"All folders in {parent_folder} have been deleted.")

    except Exception as e:
        logger.error(f"Failed to delete folders in {parent_folder}: {e}")


@callback(
    Output("download_selected", "disabled"),
    Output("download_selected_info", "children"),
    Output("download_link", "children"),
    Input("download_selected", "n_clicks"),
    Input("datatable-paging-and-sorting", "derived_viewport_selected_rows"),
    Input("datatable-paging-and-sorting", "derived_viewport_selected_row_ids"),
    State("datatable-paging-and-sorting", "derived_viewport_data"),
)
def start_download(clicks: int, rows: List, ids: List, data: List) -> tuple:
    """Start the download process for selected rows in the data table.

    Args:
        clicks (int): Number of clicks on the download button.
        rows (List): List of selected row indices.
        ids (List): List of selected row IDs.
        data (List): Data of the data table.

    Returns:
        tuple: A tuple containing:
            - bool: Whether the download button should be disabled.
            - str: Status message for the download.
            - html.A: Download link or message indicating no link available.

    """
    logger.info(f"Starting download with clicks: {clicks}, rows: {rows}, ids: {ids}.")
    triggered_id = (
        callback_context.triggered[0]["prop_id"].split(".")[0]
        if callback_context.triggered
        else None
    )
    logger.info(f"Triggered by: {triggered_id}.")

    no_href_link = html.A(
        "Download link will be available here.", href=None, style={"color": "red"}
    )

    if (rows is None) or (ids is None):
        return False, "No rows selected for download.", no_href_link

    if len(rows) > 0:
        logger.info(f"Disabled is False, Selected rows: {rows}, ids: {ids}")

        if len(rows) > 20:
            logger.info(
                f"Too many rows selected for download: {len(rows)}. Limit is 20."
            )
            return (
                True,
                "Too many rows selected for download. Limit is to 20 rows.",
                no_href_link,
            )

        if triggered_id == "download_selected":
            logger.info(f"Download button was pressed: {clicks}, rows: {rows}")

            uuid = str(uuid4())
            logger.info(f"Started download (id = {uuid}).")

            tmp_folder = app.config["TMP_FOLDER"]
            logger.info(f"Temporary folder: {tmp_folder}.")

            destination_folder = f"{tmp_folder}/download_{uuid}"
            logger.info(f"Temporary download folder: {destination_folder}.")

            selected_data = [data[i] for i in rows]
            file_paths = [data_item["system_result"] for data_item in selected_data]
            file_uuids = [data_item["system_uuid"] for data_item in selected_data]
            file_info = list(zip(file_paths, file_uuids))

            zip_file_name = f"download_{uuid}.zip"

            logger.info(f"file paths: {file_paths}")
            logger.info(f"file uuids: {file_uuids}")
            logger.info(f"destination folder: {destination_folder}")

            copy_and_zip_files(
                file_info=file_info,
                destination_folder=destination_folder,
                zip_file_name=zip_file_name,
            )

            flask_url = f"/assas_app/hdf5_download?uuid={uuid}"
            clickable_link = html.A(
                "Click here to download the zip archive",
                href=flask_url,
                target="_blank",
            )

            return True, "Zipped archive ready for download.", clickable_link

        else:
            logger.info(
                "No download button was pressed, "
                f"just selected rows: {rows}, ids: {ids}"
            )
            return True, "Press button to generate download link.", no_href_link

    else:
        logger.info(f"Disabled is True, Selected rows: {rows}, ids: {ids}")
        return True, "No rows selected for download.", no_href_link


def split_filter_part(
    filter_part: str
) -> List[str]:
    """Split a filter part into name, operator, and value.

    Args:
        filter_part (str): The filter part to split, e.g., "{name} contains 'value'".

    Returns:
        List[str]: A list containing the name, operator, and value.
            If no operator is found, returns [None, None, None].

    """
    logger.info(f"Operators: {operators}")
    logger.info(f"Filter part: {filter_part}")

    for operator_type in operators:
        for operator in operator_type:
            if operator in filter_part:
                name_part, value_part = filter_part.split(operator, 1)
                name = name_part[name_part.find("{") + 1 : name_part.rfind("}")]

                value_part = value_part.strip()
                v0 = value_part[0]

                if v0 == value_part[-1] and v0 in (""", """, "`"):
                    value = value_part[1:-1].replace("\\" + v0, v0)
                else:
                    try:
                        value = float(value_part)
                    except ValueError:
                        value = value_part

                # word operators need spaces after them in the filter string,
                # but we don't want these later
                return name, operator_type[0].strip(), value

    return [None] * 3


def apply_filters(
    dataframe: pd.DataFrame,
    filter_query: str
) -> pd.DataFrame:
    """Apply filters to the dataframe based on filter query.

    Args:
        dataframe (pd.DataFrame): The dataframe to filter
        filter_query (str): The filter query string

    Returns:
        pd.DataFrame: Filtered dataframe

    """
    if not filter_query:
        return dataframe

    filtering_expressions = filter_query.split(" && ")

    for filter_part in filtering_expressions:
        col_name, operator, filter_value = split_filter_part(filter_part)

        if col_name is None or operator is None:
            continue

        try:
            if operator == "eq":
                dataframe = dataframe.loc[dataframe[col_name] == filter_value]
            elif operator == "ne":
                dataframe = dataframe.loc[dataframe[col_name] != filter_value]
            elif operator == "lt":
                dataframe = dataframe.loc[dataframe[col_name] < filter_value]
            elif operator == "le":
                dataframe = dataframe.loc[dataframe[col_name] <= filter_value]
            elif operator == "gt":
                dataframe = dataframe.loc[dataframe[col_name] > filter_value]
            elif operator == "ge":
                dataframe = dataframe.loc[dataframe[col_name] >= filter_value]
            elif operator == "contains":
                if col_name in dataframe.columns:
                    # Convert to string for contains operation
                    dataframe = dataframe.loc[
                        dataframe[col_name]
                        .astype(str)
                        .str.contains(str(filter_value), case=False, na=False)
                    ]
            elif operator == "datestartswith":
                if col_name in dataframe.columns:
                    dataframe = dataframe.loc[
                        dataframe[col_name]
                        .astype(str)
                        .str.startswith(str(filter_value))
                    ]
        except Exception as e:
            logger.warning(f"Error applying filter {filter_part}: {e}")
            continue

    return dataframe


@callback(
    [
        Output("datatable-paging-and-sorting", "data"),
        Output("pagination-contents", "children"),
    ],
    [
        Input("pagination", "active_page"),
        Input("datatable-page-size", "value"),
        Input("datatable-paging-and-sorting", "sort_by"),
        Input("datatable-paging-and-sorting", "filter_query"),
        Input("reload_page", "n_clicks"),
    ],
    prevent_initial_call=False,
)
def update_table_with_pagination(
    active_page: int | None = None,
    page_size: int | None = None,
    sort_by: List[dict] | None = None,
    filter_query: str | None = None,
    n_clicks: int | None = None,
)-> Tuple[List[dict], str]:
    """Update table data based on pagination, sorting, filtering, and refresh.

    Args:
        active_page (int | None): Current active page number. Defaults to 1.
        page_size (int | None): Number of rows per page. Defaults to PAGE_SIZE.
        sort_by (List[dict] | None): List of columns to sort by.
        filter_query (str | None): Filter query string to apply.
        n_clicks (int | None): Number of clicks on the reload button.

    Returns:
        Tuple[List[dict], str]: A tuple containing:
            - List of dictionaries representing the paginated data.
            - Pagination info text.

    """
    # Use default values if None
    if active_page is None:
        active_page = 1
    if page_size is None:
        page_size = PAGE_SIZE
    if sort_by is None:
        sort_by = []

    # Get fresh data (especially important after refresh)
    dataframe = update_table_data()
    logger.info(f"Loaded {len(dataframe)} records from database")

    # Apply filtering using enhanced function
    if filter_query:
        logger.info(f"Applying filter: {filter_query}")
        dataframe = apply_filters(dataframe, filter_query)
        logger.info(f"After filtering: {len(dataframe)} records")

    # Apply sorting
    if len(sort_by):
        logger.info(f"Applying sort: {sort_by}")
        try:
            dataframe = dataframe.sort_values(
                [col["column_id"] for col in sort_by],
                ascending=[col["direction"] == "asc" for col in sort_by],
                inplace=False,
            )
            logger.info("Sorting applied successfully")
        except Exception as e:
            logger.warning(f"Error applying sort: {e}")

    # Calculate pagination
    total_records = len(dataframe)
    start_idx = (active_page - 1) * page_size
    end_idx = start_idx + page_size

    # Get paginated data
    paginated_data = dataframe.iloc[start_idx:end_idx].to_dict("records")

    # Update pagination info text
    start_record = start_idx + 1 if total_records > 0 else 0
    end_record = min(end_idx, total_records)

    pagination_text = (
        f"Page {active_page} | Showing {start_record}-{end_record} of {total_records}"
    )

    if filter_query:
        pagination_text += " (filtered)"

    logger.info(f"Pagination: {pagination_text}")

    return paginated_data, pagination_text


@callback(
    Output("pagination", "max_value"),
    [
        Input("datatable-page-size", "value"),
        Input("datatable-paging-and-sorting", "filter_query"),
        Input("reload_page", "n_clicks"),
    ],
    prevent_initial_call=False,
)
def update_pagination_max_value(
    page_size: int | None = None,
    filter_query: str | None = None,
    n_clicks: int | None = None,
) -> int:
    """Update pagination max value when page size, filter, or data changes.

    Args:
        page_size (int | None): Number of rows per page. If None, uses default
            PAGE_SIZE.
        filter_query (str | None): Filter query string to apply.
        n_clicks (int | None): Number of clicks on the reload button.

    Returns:
        int: Total number of pages based on current data and filters.

    """
    if page_size is None:
        page_size = PAGE_SIZE

    # Get fresh data
    dataframe = update_table_data()

    # Apply filtering to get accurate count
    if filter_query:
        dataframe = apply_filters(dataframe, filter_query)

    total_pages = math.ceil(len(dataframe) / page_size) if len(dataframe) > 0 else 1
    logger.info(f"Updated pagination: {total_pages} total pages")

    return total_pages


@callback(
    Output("pagination", "active_page"),
    [
        Input("datatable-page-size", "value"),
        Input("datatable-paging-and-sorting", "filter_query"),
    ],
    prevent_initial_call=True,
)
def reset_pagination_on_changes(
    page_size: int,
    filter_query: str,
) -> int:
    """Reset to page 1 when page size or filter changes.

    Args:
        page_size (int): Number of rows per page.
        filter_query (str): Filter query string to apply.

    Returns:
        int: Always returns 1 to reset pagination to the first page.

    """
    return 1


@callback(
    Output("datatable-paging-and-sorting", "style_table"),
    Input("datatable-page-size", "value"),
    prevent_initial_call=False,
)
def update_table_height(
    page_size: int | None = None,
)-> dict:
    """Dynamically adjust table height based on page size.

    Args:
        page_size (int | None): Number of rows per page.
        If None, uses default PAGE_SIZE.

    Returns:
        dict: CSS styles for the data table container.

    """
    if page_size is None:
        page_size = PAGE_SIZE

    # Calculate height based on number of rows
    padding_height = 40  # Additional padding and borders

    # Calculate total height needed
    calculated_height = (page_size * 60) + 50 + padding_height

    # Convert to CSS height
    table_height = f"{calculated_height}px"

    return {
        "overflowX": "auto",
        "overflowY": "visible",
        "borderRadius": "12px",
        "border": "2px solid #e0e6ed",
        "fontFamily": "Arial, sans-serif",
        "width": "100%",
        "height": table_height,
        "minHeight": "300px",
        "maxHeight": "none",
        "boxShadow": "0 4px 12px rgba(0, 0, 0, 0.15)",
        "tableLayout": "fixed",
    }


@callback(
    Output("collapse-usage-guide", "is_open"),
    Input("toggle-usage-guide", "n_clicks"),
    prevent_initial_call=True,
)
def toggle_usage_guide(
    n_clicks: int
) -> bool:
    """Toggle the visibility of the usage guide section.

    Args:
        n_clicks (int): Number of clicks on the toggle button.

    Returns:
        bool: True if the section should be open, False if it should be closed.

    """
    logger.info(f"Usage guide toggle clicked {n_clicks} times.")
    return n_clicks % 2 == 1  # Toggle between open and closed


@callback(
    Output("mobile-current-page", "children"),
    Input("pagination", "active_page"),
    prevent_initial_call=False,
)
def update_mobile_page_display(
    active_page: int | None = None,
)-> str:
    """Update mobile pagination display.

    Args:
        active_page (int | None): Current active page number.

    Returns:
        str: Current active page number as a string.

    """
    if active_page is None:
        active_page = 1
    return str(active_page)


@callback(
    [
        Output("pagination", "active_page", allow_duplicate=True),
        Output("mobile-prev-btn", "disabled"),
        Output("mobile-next-btn", "disabled"),
    ],
    [Input("mobile-prev-btn", "n_clicks"), Input("mobile-next-btn", "n_clicks")],
    [State("pagination", "active_page"), State("pagination", "max_value")],
    prevent_initial_call=True,
)
def handle_mobile_navigation(
    prev_clicks: int,
    next_clicks: int,
    current_page: int,
    max_pages: int,
) -> Tuple[int, bool, bool]:
    """Handle mobile pagination navigation.

    Args:
        prev_clicks (int): Number of clicks on the previous button.
        next_clicks (int): Number of clicks on the next button.
        current_page (int): Current active page number.
        max_pages (int): Maximum number of pages available.

    Returns:
        Tuple[int, bool, bool]: A tuple containing:
            - New active page number.
            - Whether the previous button should be disabled.
            - Whether the next button should be disabled.

    """
    triggered = (
        callback_context.triggered[0]["prop_id"] if callback_context.triggered else None
    )

    if current_page is None:
        current_page = 1
    if max_pages is None:
        max_pages = 1

    new_page = current_page

    if triggered == "mobile-prev-btn.n_clicks" and prev_clicks:
        new_page = max(1, current_page - 1)
    elif triggered == "mobile-next-btn.n_clicks" and next_clicks:
        new_page = min(max_pages, current_page + 1)

    # Determine button states
    prev_disabled = new_page <= 1
    next_disabled = new_page >= max_pages

    return new_page, prev_disabled, next_disabled

# Add these callback functions for export functionality

def prepare_export_data(filter_query, sort_by, export_options):
    """Prepare data for export with filters and sorting applied."""
    try:
        # Get fresh data
        df = update_table_data()
        
        # Apply filters if requested
        if "apply_filters" in export_options and filter_query:
            df = apply_filters(df, filter_query)
        
        # Apply sorting if requested
        if sort_by and len(sort_by) > 0:
            df = df.sort_values(
                [col["column_id"] for col in sort_by],
                ascending=[col["direction"] == "asc" for col in sort_by],
                inplace=False,
            )
        
        return df
    except Exception as e:
        logger.error(f"Error preparing export data: {e}")
        return pd.DataFrame()  # Return empty DataFrame on error

def clean_data_for_export(df, export_options):
    """Clean data for export by removing HTML and preparing for CSV/Excel."""
    try:
        df_clean = df.copy()
        
        # Remove HTML from specific columns
        html_columns = ['meta_name', 'system_status', 'system_download']
        
        for col in html_columns:
            if col in df_clean.columns:
                if col == 'meta_name':
                    # Extract text from links for dataset names
                    df_clean[col] = df_clean[col].astype(str).str.extract(r'>([^<]+)<', expand=False).fillna(df_clean[col])
                elif col == 'system_status':
                    # Extract status text from HTML spans
                    df_clean[col] = df_clean[col].astype(str).str.extract(r'>([^<]+)<', expand=False).fillna(df_clean[col])
                elif col == 'system_download':
                    if "include_links" not in export_options:
                        # Replace download links with simple text
                        df_clean[col] = df_clean[col].apply(lambda x: 
                            "hdf5 file available" if "hdf5 file" in str(x) and "no-download" not in str(x)
                            else "no hdf5 file"
                        )
        
        # Clean column names for better readability
        column_mapping = {
            'system_index': 'Index',
            'meta_name': 'Dataset Name',
            'system_status': 'Status',
            'system_date': 'Upload Date',
            'system_user': 'User',
            'system_size': 'Binary Size',
            'system_size_hdf5': 'HDF5 Size',
            'system_download': 'Download Status',
            'system_uuid': 'UUID'
        }
        
        # Rename columns that exist in the dataframe
        existing_columns = {k: v for k, v in column_mapping.items() if k in df_clean.columns}
        df_clean = df_clean.rename(columns=existing_columns)
        
        return df_clean
    except Exception as e:
        logger.error(f"Error cleaning export data: {e}")
        return df  # Return original DataFrame on error

def format_excel_export_basic(worksheet, df):
    """Apply basic formatting to Excel export."""
    try:
        from openpyxl.styles import Font, PatternFill
        
        # Header formatting
        header_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Apply header formatting
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    except Exception as e:
        logger.error(f"Error formatting Excel export: {e}")

def add_metadata_sheet_basic(writer, filter_query, sort_by, export_options):
    """Add basic metadata sheet to Excel export."""
    try:
        # Create metadata information
        metadata_info = [
            ["Export Information", ""],
            ["Export Date", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Export Source", "ASSAS Database Training Dataset Index"],
            ["Applied Filters", filter_query if filter_query else "None"],
            ["Applied Sorting", ", ".join([f"{col['column_id']} ({col['direction']})" for col in (sort_by or [])]) if sort_by else "None"],
            ["Export Options", ", ".join(export_options) if export_options else "Default"],
        ]
        
        # Convert to DataFrame and add to Excel
        metadata_df = pd.DataFrame(metadata_info, columns=['Field', 'Value'])
        metadata_df.to_excel(writer, sheet_name='Export Info', index=False)
    except Exception as e:
               logger.error(f"Error adding metadata sheet: {e}")


@callback(
    [Output("download-csv-database", "data"),
     Output("export-status-database", "children")],
    Input("export-csv-btn", "n_clicks"),
    [State("datatable-paging-and-sorting", "derived_viewport_data"),
     State("datatable-paging-and-sorting", "data"), 
     State("datatable-paging-and-sorting", "filter_query"),
     State("datatable-paging-and-sorting", "sort_by"),
     State("export-options", "value")],
    prevent_initial_call=True
)
def export_csv(n_clicks, current_page_data, all_data, filter_query, sort_by, export_options):


    if not n_clicks:
        raise PreventUpdate
    
    try:
        export_options = export_options or []
        
        # Determine what data to export
        if "current_page" in export_options:
            export_data = current_page_data or []
            status_msg = f"✅ Exported {len(export_data)} records (current page) to CSV"
        else:
            # Get all data with current filters and sorting
            df = prepare_export_data(filter_query, sort_by, export_options)
            export_data = df.to_dict('records')
            status_msg = f"✅ Exported {len(export_data)} records to CSV"
        
        if not export_data:
            logger.warning("No data to export")
            return dash.no_update, "⚠️ No data to export"
        
        # Convert to DataFrame for CSV export
        df_export = pd.DataFrame(export_data)
        
        # Clean data for export
        df_export = clean_data_for_export(df_export, export_options)
        
        # Generate timestamp for filename
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        filename = f"assas_database_export_{timestamp}.csv"
        
        # Create CSV download
        csv_string = df_export.to_csv(index=False, encoding='utf-8')
        
        return dict(content=csv_string, filename=filename), status_msg
        
    except Exception as e:
        logger.error(f"CSV export error: {e}")
        return dash.no_update, f"❌ Export failed: {str(e)}"
    
@callback(
    [Output("download-excel-database", "data"),
     Output("export-status-database", "children", allow_duplicate=True)],
    Input("export-excel-btn", "n_clicks"),
    [State("datatable-paging-and-sorting", "derived_viewport_data"),
     State("datatable-paging-and-sorting", "data"),
     State("datatable-paging-and-sorting", "filter_query"),
     State("datatable-paging-and-sorting", "sort_by"),
     State("export-options", "value")],
    prevent_initial_call=True
)
def export_excel(n_clicks, current_page_data, all_data, filter_query, sort_by, export_options):
    """Export table data to Excel format."""
    if not n_clicks:
        raise PreventUpdate
    
    try:
        export_options = export_options or []
        
        # Determine what data to export
        if "current_page" in export_options:
            export_data = current_page_data or []
            status_msg = f"✅ Exported {len(export_data)} records (current page) to Excel"
        else:
            # Get all data with current filters and sorting
            df = prepare_export_data(filter_query, sort_by, export_options)
            export_data = df.to_dict('records')
            status_msg = f"✅ Exported {len(export_data)} records to Excel"
        
        if not export_data:
            return dash.no_update, "⚠️ No data to export"
        
        # Convert to DataFrame for Excel export
        df_export = pd.DataFrame(export_data)
        
        # Clean data for export
        df_export = clean_data_for_export(df_export, export_options)
        
        # Generate timestamp for filename
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        filename = f"assas_database_export_{timestamp}.xlsx"
        
        # 🔧 FIX: Create Excel file in memory with proper encoding
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Main data sheet
            df_export.to_excel(writer, sheet_name='ASSAS Database', index=False)
            
            # Get workbook and worksheet for formatting
            workbook = writer.book
            worksheet = workbook['ASSAS Database']
            
            # Apply basic formatting
            format_excel_export_basic(worksheet, df_export)
            
            # Add metadata sheet
            add_metadata_sheet_basic(writer, filter_query, sort_by, export_options)
        
        # 🔧 FIX: Encode binary data to base64 for JSON serialization
        buffer.seek(0)
        excel_data = buffer.getvalue()
        
        # Encode to base64 string for JSON serialization
        encoded_data = base64.b64encode(excel_data).decode('utf-8')
        
        return {
            "content": encoded_data,
            "filename": filename,
            "base64": True,  # Tell Dash this is base64 encoded
            "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }, status_msg
        
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        return dash.no_update, f"❌ Export failed: {str(e)}"

